import os
import csv
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from statistics import stdev
import openpyxl
from openpyxl import load_workbook

#Folder path containing all folders of csv
folder_path = 'C:\\Users\\mablu\\OneDrive\\Documents\\Senior Design\\600 RPM Load Testing'


#sensor = 'Analog - 1-25-2024 12-37-30.522 PM.csv'
#encoder = 'Counter - 1-25-2024 12-37-30.5 PM.csv'


# define encoder sensor parameters
delta_t = 0.1 # inverse of the encoder sample rate configured in DAQami
tick_count = 2048 # number of ticks in one revolution
revs_per_tick = 1 / tick_count # revolutions per tick
gear_ratio = 3

def parse_csv_file(file_path):
    with open(file_path, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            # Process each row as needed
            #print(row)
            return

def process_folder(folder_path):
    file_appendix = []
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            if file_name.endswith('.csv') and 'a' in file_name:
                file_path = os.path.join(root, file_name)
                print(f'Parsing CSV file: {file_path}')
                Analog_data = file_path
                file_appendix.append(Analog_data)
                parse_csv_file(file_path)
            elif file_name.endswith('.csv') and 'c' in file_name:
                file_path = os.path.join(root, file_name)
                print(f'Parsing CSV file: {file_path}')
                Encoder = file_path
                parse_csv_file(file_path)
                file_appendix.append(Encoder)
    return file_appendix            

def parse_csv(file_path):
    # Initialize empty dictionary to store column vectors
    column_vectors = {}

    # Open the CSV file
    with open(file_path, 'r') as csv_file:
        # Create a CSV reader
        reader = csv.reader(csv_file)

        # Read the first 6 lines (assuming headers are on the 7th line)
        for _ in range(6):
            next(reader)

        # Read the headers from the 7th line
        headers = next(reader)

        # Initialize empty lists for each column
        for header in headers:
            column_vectors[header] = []

        # Read the rest of the file and populate column vectors
        for row in reader:
            for header, value in zip(headers, row):
                column_vectors[header].append(value)

    return column_vectors

def convert_encoder_to_rpm(encoder_voltage):
    measured_L_rpm = [] # intializing measured rpm list, L corresponds with low speed shaft reading
    # converts ticks to Low Speed Shaft rpm given a time step and tick count. Excludes final data point in raw_encoder list
    for i in range(0, len(encoder_voltage) - 1):
        test_value = round(((encoder_voltage[i+1] - encoder_voltage[i])/delta_t) * revs_per_tick * 60 * gear_ratio, 4) # also converts from rps to rpm
        measured_L_rpm.append(test_value)
    return measured_L_rpm

def convert_torque(raw_torque):
    measured_torque = []
    for i in range(len(raw_torque)):
        if raw_torque[i][0] == '-':
            sensed_value = round(float(raw_torque[i][1:]), 4)
        else:
            sensed_value = round(float(raw_torque[i]), 4)
        measured_torque.append(round(sensed_value * 10 / gear_ratio, 2)) # conversion equation: (V_sensed[V] * 10[Nm/V]) = Torque[Nm]
    return measured_torque

def convert_temp(raw_temp):
    measured_temp = []
    for i in range(len(raw_temp)):
        sensed_value = round(float(raw_temp[i]), 2)
        actual_value = (sensed_value) * 100 - 273.15 # conversion equation: @@@@@@@@@@@
        measured_temp.append(actual_value)

    return measured_temp

def convert_voltage(raw_voltage):
    measured_voltage = []
    for i in range(len(raw_voltage)):
        sensed_value = round(float(raw_voltage[i]), 2)
        actual_value = 3.125 * (sensed_value / 250) * 1000 - 12.5 # conversion equation: 3.125[V/mA] * (V_sensed / R) * 10^3 - 12.5V = Voltage[V]
        measured_voltage.append(actual_value)
    return measured_voltage
    
def convert_current(raw_current):
    measured_current = []
    for i in range(len(raw_current)):
        sensed_value = round(float(raw_current[i]), 2)
        actual_value = (3.125 * (sensed_value / 250) * 1000 - 12.5) # conversion equation: 3.125[A/mA] * (V_sensed[V] / R[ohms]) * 10^3 - 12.5[A] = Amperage[A]
        measured_current.append(actual_value)
    return measured_current

def get_input_power(measured_torque, measured_L_rpm, sample_count): 
    time = []
    input_power = []

    # assigning time row vector
    sample_rate = 1000 # samples per second
    sample_step = 1 / sample_rate
    for i in range(0, sample_count):
        time.append(i * sample_step)

    # encoder samples at different rates, have to correlate times with sensor data
    time_enc = []
    sample_rate_enc = 10 # samples per second
    sample_step_enc = 1 / sample_rate_enc
    sample_count_enc = len(measured_L_rpm) # determine how many samples were collected
    for i in range(0, sample_count_enc):
        time_enc.append(round(i * sample_step_enc, 1))

    rpm_time_series = np.array([time_enc, measured_L_rpm])
    torque_time_series = np.array([time, measured_torque])



    for i in range(0, len(rpm_time_series[0])):
        index = np.where(i == torque_time_series[0])[0]
        input_power.append(torque_time_series[1,index] * rpm_time_series[1,i] * (3.14159 / 30))

    return input_power

def get_output_power(measured_current, measured_voltage, sample_count): 
    output_power = []

    for i in range(0, sample_count):
        i_power = measured_current[i] * measured_voltage[i]
        output_power.append(i_power)
    return output_power

def statistics_calc(data):
    avg_val = round(sum(data) / len(data), 3)
    max_val = round(max(data), 3)
    min_val = round(min(data), 3)
    std_dev = round(stdev(data), 3)

    data = {
        'Average' : avg_val, 
        'Max' : max_val,
        'Min' : min_val,
        'STDEV' : std_dev,
    }

    return data

def get_stats(data):
    Torque = statistics_calc(data.get('Torque'))
    Voltage = statistics_calc(data.get('Voltage'))
    Current  = statistics_calc(data.get('Current'))
    Temp1 = statistics_calc(data.get('Temp1'))
    Temp2 = statistics_calc(data.get('Temp2'))
    Encoder = statistics_calc(data.get('Measure Low RPM'))
    #InputPower = statistics_calc(data.get('Input Power'))
    #OutputPower = statistics_calc(data.get('Output Power'))
    data = {
        'Torque' : Torque,
        'Temp1' : Temp1,
        'Temp2' : Temp2,
        'Voltage' : Voltage,
        'Current' : Current,
        'Measure Low RPM' : Encoder
        #'Input Power' : InputPower,
        #'Output Power' : OutputPower
    }
    return data

def convert_data(file1, file2):
    result = parse_csv(file1)
    encoder = parse_csv(file2)
    raw_torque = result.get('Torque (V)')
    raw_temp1 = result.get('Temp1 (V)')
    raw_temp2 = result.get('Temp2 (V)')
    raw_voltage = result.get('CH2H Volt (V)')
    raw_current = result.get('CH3H Current (V)')
    raw_encoder = encoder.get('Encoder (Ticks)')
    raw_encoder = [int(i) for i in raw_encoder]

    time = []
    # assigning time row vector
    sample_rate = 100 # samples per second
    sample_step = 1 / sample_rate
    sample_count = len(raw_torque)
    for i in range(0, sample_count):
        time.append(i * sample_step)



    measured_torque = convert_torque(raw_torque)
    measured_temp1 = convert_temp(raw_temp1)
    measured_temp2 = convert_temp(raw_temp2)
    measured_voltage = convert_voltage(raw_voltage)
    measured_current = convert_current(raw_current)
    measured_L_rpm = convert_encoder_to_rpm(raw_encoder)
    input_power = get_input_power(measured_torque, measured_L_rpm, sample_count)
    output_power = get_output_power(measured_current, measured_voltage, sample_count)
    
    final_data = {
        'Torque' : measured_torque,
        'Temp1' : measured_temp1,
        'Temp2' : measured_temp2,
        'Voltage' : measured_voltage,
        'Current' : measured_current,
        'Measure Low RPM' : measured_L_rpm,
        'Time' :time, 
        'Input Power' : input_power, 
        'Output Power' : output_power
        }
    
    return final_data

def average_every_n_values(data, n):
    result = []
    for i in range(0, len(data), n):
        chunk = data[i:i + n]
        if chunk:  # Ensure the chunk is not empty
            result.append(sum(chunk) / len(chunk))
    return result

def export_to_excel(Stats, data):
    input_power_avg = Stats.get('Torque').get('Average') * Stats.get('Measure Low RPM').get('Average') * 3.14159 / 30 # Nm and radians per second
    output_power_avg =  Stats.get('Current').get('Average') *  Stats.get('Voltage').get('Average')
    efficiency_avg_percentage = round(output_power_avg / input_power_avg * 100, 2)


    Torque = average_every_n_values(data.get('Torque'), 1000)
    Time =   average_every_n_values(data.get('Time'), 1000)
    Voltage = average_every_n_values(data.get('Voltage'), 1000)
    Current = average_every_n_values(data.get('Current'), 1000)
    Temp1 = average_every_n_values(data.get('Temp1'), 1000)
    Temp2 = average_every_n_values(data.get('Temp2'), 1000)

    

    df = pd.DataFrame(Stats)
    
    # Specify the Excel file path
    excel_file_path = 'Nacelle_Output.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    last_row = sheet.max_row + 2
    print("Last row is ", last_row)
    with pd.ExcelWriter('Nacelle_Output.xlsx', engine='openpyxl') as writer:
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet_name = 'df', index = False)
        writer.save()
    
    print(f"Excel table created and saved to: {excel_file_path}")

def controller(data):
    Stats = get_stats(data)
    Torque = average_every_n_values(data.get('Torque'), 1000)
    Voltage = average_every_n_values(data.get('Voltage'), 1000)
    Current = average_every_n_values(data.get('Current'), 1000)
    Temp1 = average_every_n_values(data.get('Temp1'), 1000)
    Temp2 = average_every_n_values(data.get('Temp2'), 1000)
    Time =   average_every_n_values(data.get('Time'), 1000)


    input_power_avg = Stats.get('Torque').get('Average') * Stats.get('Measure Low RPM').get('Average') * 3.14159 / 30 # Nm and radians per second
    output_power_avg =  Stats.get('Current').get('Average') *  Stats.get('Voltage').get('Average')
    efficiency_avg_percentage = round(output_power_avg / input_power_avg * 100, 2)

    power_data = [
        ["Input Power (W)" , input_power_avg], 
        ["Output Power (W)" , output_power_avg], 
        ["Efficiency" , efficiency_avg_percentage]
    ]

    power_output = pd.DataFrame(power_data)


    def plot_torque():
        plt.plot(Time, Torque, marker='o', linestyle='-', color='b', label='Line Graph')
        step_size = 0.1
        plt.yticks(np.arange(min(Torque) - 0.1, max(Torque) + 0.1, step_size))
        # Adding labels and title
        plt.xlabel('Time')
        plt.ylabel('Torque')
        plt.title('Torque over Time')
        #plt.show()

    def plot_voltage():
        plt.plot(Time, Voltage, marker='o', linestyle='-', color='b', label='Line Graph')
        # Adding labels and title
        plt.xlabel('Time')
        plt.ylabel('Voltage')
        plt.title('Voltage over Time')
        #plt.show()

    def plot_current():
        plt.plot(Time, Current, marker='o', linestyle='-', color='b', label='Line Graph')
        # Adding labels and title
        plt.xlabel('Time')
        plt.ylabel('Current')
        plt.title('Current over Time')
        #plt.show()

    def plot_temp1():
        plt.plot(Time, Temp1, marker='o', linestyle='-', color='b', label='Line Graph')
        # Adding labels and title
        plt.xlabel('Time')
        plt.ylabel('Temp1')
        plt.title('Temp1 over Time')
        #plt.show()

    def plot_temp2():
        plt.plot(Time, Temp2, marker='o', linestyle='-', color='b', label='Line Graph')
        # Adding labels and title
        plt.xlabel('Time')
        plt.ylabel('Temp2')
        plt.title('Temp2 over Time')
        #plt.show()

    #plot_torque()
    #plot_voltage()
    #plot_current()
    #plot_temp1()
    #plot_temp2()
    #export_to_excel(Stats, data)
    df = pd.DataFrame(Stats)
    return(df, power_output)

def main(File_paths):
    print(File_paths)
    appendix_dict = []
    i=0
    while i < len(File_paths):
        sensor = File_paths[i]
        print("The Sensor csv is " , sensor)
        i+=1
        encoder = File_paths[i]
        print("The Encoder csv is " , encoder)
        i+=1
        appendix_dict.append([sensor, encoder])

    Averaged_data_per_test = []
    final_output_power = []
    for x in appendix_dict:
        data = convert_data(x[0], x[1])
        print(get_stats(data))
        Averaged_data_per_test.append(controller(data)[0])
        final_output_power.append(controller(data)[1])

    df = pd.DataFrame()
    for x in Averaged_data_per_test:
        df = df.append(x)

    power = pd.DataFrame()
    for x in final_output_power:
        power = power.append(x)

    


    excel_file_path_power = 'Power_Data.xlsx'
    excel_file_path_sensor = 'Sensor_Data.xlsx'
    df.to_excel(excel_file_path_sensor, header = True, index=True, sheet_name="Sensor Data")
    power.to_excel(excel_file_path_power, header = False, index=False , sheet_name="Power Data")
    
appendix = process_folder(folder_path)
main(appendix)